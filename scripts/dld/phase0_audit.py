#!/usr/bin/env python3
"""Build the local, read-only DLD Phase 0 inventory and audit.

The source directory is only opened for reading. Raw rows are materialized into
an ignored local DuckDB database; committed reports contain counts and metadata,
never source records.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import duckdb
from openpyxl import load_workbook


DATASET_TABLES = {
    "Accredited Trainers by the Dubai Real Estate Institute": "accredited_trainers",
    "Approved Escrow Account Agents": "escrow_agents",
    "Building and Property Project Records": "buildings",
    "Developers Recorded in Dubai Land Department": "developers",
    "Elders and People of Determination": "elders_people_of_determination",
    "Free Zone Companies Licensing": "free_zone_companies",
    "Land Registry": "land_registry",
    "Licenced Owner Associations": "owner_associations",
    "Licensed Real Estate Valuators": "valuators",
    "Lookup Dubai Community Areas": "areas",
    "Lookup Real Estate Market Types": "market_types",
    "Lookup Real Estate Transactions Groups": "transaction_groups",
    "Lookup Real Estate Transactions Procedures": "transaction_procedures",
    "Owners Association Service Charges": "service_charges",
    "Property Map Requests": "map_requests",
    "Property Valuation Records": "valuations",
    "Real Estate Brokers": "brokers",
    "Real Estate Licenses": "licences",
    "Real Estate Offices": "offices",
    "Real Estate Permits": "permits",
    "Real Estate Projects": "projects",
    "Real Estate Transactions": "transactions",
    "Registered Freehold Real Estate Units": "units",
    "Rent Contracts": "rent_contracts",
    "Residential Sale Index": "residential_sale_index",
}


RELATIONSHIPS = [
    ("projects.area_id -> areas.area_id", "projects", "area_id", "areas", "area_id"),
    ("transactions.area_id -> areas.area_id", "transactions", "area_id", "areas", "area_id"),
    ("rent_contracts.area_id -> areas.area_id", "rent_contracts", "area_id", "areas", "area_id"),
    ("units.area_id -> areas.area_id", "units", "area_id", "areas", "area_id"),
    ("land_registry.area_id -> areas.area_id", "land_registry", "area_id", "areas", "area_id"),
    ("valuations.area_id -> areas.area_id", "valuations", "area_id", "areas", "area_id"),
    ("buildings.area_id -> areas.area_id", "buildings", "area_id", "areas", "area_id"),
    ("projects.developer_id -> developers.developer_id", "projects", "developer_id", "developers", "developer_id"),
    ("projects.master_developer_id -> developers.developer_id", "projects", "master_developer_id", "developers", "developer_id"),
    ("projects.escrow_agent_id -> escrow_agents.escrow_agent_number", "projects", "escrow_agent_id", "escrow_agents", "escrow_agent_number"),
    ("buildings.project_id -> projects.project_id", "buildings", "project_id", "projects", "project_id"),
    ("units.project_id -> projects.project_id", "units", "project_id", "projects", "project_id"),
    ("land_registry.project_id -> projects.project_id", "land_registry", "project_id", "projects", "project_id"),
    ("service_charges.project_id -> projects.project_id", "service_charges", "project_id", "projects", "project_id"),
    ("brokers.real_estate_id -> offices.real_estate_id", "brokers", "real_estate_id", "offices", "real_estate_id"),
    ("brokers.real_estate_number -> offices.real_estate_number", "brokers", "real_estate_number", "offices", "real_estate_number"),
    ("offices.participant_id -> licences.participant_id", "offices", "participant_id", "licences", "participant_id"),
    ("offices.license_number -> licences.license_number", "offices", "license_number", "licences", "license_number"),
    ("developers.participant_id -> licences.participant_id", "developers", "participant_id", "licences", "participant_id"),
    ("developers.license_number -> licences.license_number", "developers", "license_number", "licences", "license_number"),
    ("permits.license_number -> licences.license_number", "permits", "license_number", "licences", "license_number"),
    ("permits.license_number -> offices.license_number", "permits", "license_number", "offices", "license_number"),
]


DATE_NAME_RE = re.compile(r"(?:^|_)(?:date|year|month|timestamp)(?:_|$)|(?:_date$|_year$|_month$)", re.I)
IDENTIFIER_RE = re.compile(r"(?:^|_)(?:id|number|no|code)$|_(?:id|number|no|code)$", re.I)
INTERNAL_EXACT = {
    "load_timestamp", "participant_id", "property_id", "parent_property_id",
    "grandparent_property_id", "parcel_id", "contract_id", "transaction_id",
    "application_id", "request_id", "procedure_number", "pre_registration_number",
    "rent_contract_no", "parent_parmits_id", "parent_permits_id", "main_office_id",
}
INTERNAL_CONTACT = {"phone", "fax", "email", "webpage"}
GRANULAR_PROPERTY_DATASETS = {
    "Land Registry", "Property Valuation Records", "Real Estate Transactions",
    "Registered Freehold Real Estate Units", "Rent Contracts",
}


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def qident(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def qtext(value: str | Path) -> str:
    return "'" + str(value).replace("'", "''") + "'"


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        fieldnames = list(rows[0]) if rows else []
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def sha256_file(path: Path, chunk_size: int = 8 * 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest()


def source_files(source: Path) -> list[Path]:
    return sorted((p for p in source.rglob("*") if p.is_file()), key=lambda p: p.as_posix().lower())


def inventory_source(source: Path) -> list[dict[str, Any]]:
    rows = []
    for index, path in enumerate(source_files(source), start=1):
        stat = path.stat()
        print(f"  hashing {index:02d}/62 {path.relative_to(source)}", flush=True)
        rows.append({
            "relative_path": path.relative_to(source).as_posix(),
            "dataset": path.parent.name,
            "extension": path.suffix.lower(),
            "bytes": stat.st_size,
            "modified_ns": stat.st_mtime_ns,
            "sha256": sha256_file(path),
        })
    return rows


def read_dictionaries(source: Path) -> tuple[list[dict[str, Any]], dict[str, list[str]]]:
    rows: list[dict[str, Any]] = []
    declared_keys: dict[str, list[str]] = defaultdict(list)
    for path in sorted(source.rglob("*.xlsx")):
        dataset = path.parent.name
        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            values = sheet.iter_rows(values_only=True)
            headers = [str(value).strip() if value is not None else "" for value in next(values)]
            expected = ["Name", "Description", "Primary Key", "Size", "Data Type", "Range of values", "Classification", "Language", "Foreign key"]
            if headers != expected:
                raise ValueError(f"Unexpected dictionary header in {path}: {headers}")
            for ordinal, values_row in enumerate(values, start=1):
                if not any(value is not None for value in values_row):
                    continue
                record = dict(zip(headers, values_row))
                field = str(record["Name"]).strip().lower()
                is_pk = str(record["Primary Key"] or "").strip().lower() == "yes"
                if is_pk:
                    declared_keys[dataset].append(field)
                rows.append({
                    "dataset": dataset,
                    "workbook": path.relative_to(source).as_posix(),
                    "sheet": sheet.title,
                    "ordinal": ordinal,
                    "field": field,
                    "description": str(record["Description"] or "").strip(),
                    "declared_primary_key": is_pk,
                    "declared_size": str(record["Size"] or "").strip(),
                    "declared_data_type": str(record["Data Type"] or "").strip(),
                    "range_of_values": str(record["Range of values"] or "").strip(),
                    "source_classification": str(record["Classification"] or "").strip(),
                    "language": str(record["Language"] or "").strip(),
                    "declared_foreign_key": str(record["Foreign key"] or "").strip().lower() == "yes",
                })
        workbook.close()
    return rows, dict(declared_keys)


def group_csvs(source: Path) -> dict[str, list[Path]]:
    grouped: dict[str, list[Path]] = defaultdict(list)
    for path in source.rglob("*.csv"):
        grouped[path.parent.name].append(path)
    return {dataset: sorted(paths) for dataset, paths in grouped.items()}


def connect_database(path: Path) -> duckdb.DuckDBPyConnection:
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = duckdb.connect(str(path))
    conn.execute("SET threads = 8")
    conn.execute("SET preserve_insertion_order = false")
    conn.execute("SET memory_limit = '12GB'")
    conn.execute(f"SET temp_directory = {qtext(path.parent / 'tmp')}")
    return conn


def materialize(conn: duckdb.DuckDBPyConnection, table: str, paths: list[Path], reuse: bool) -> None:
    exists = conn.execute("SELECT count(*) FROM information_schema.tables WHERE table_name = ?", [table]).fetchone()[0]
    if reuse and exists:
        return
    path_list = "[" + ",".join(qtext(path) for path in paths) + "]"
    conn.execute(f"DROP TABLE IF EXISTS {qident(table)}")
    conn.execute(
        f"CREATE TABLE {qident(table)} AS "
        f"SELECT * EXCLUDE(filename), filename AS _source_file "
        f"FROM read_csv({path_list}, header=true, all_varchar=true, union_by_name=true, "
        f"filename=true, strict_mode=true, null_padding=false)"
    )


def table_columns(conn: duckdb.DuckDBPyConnection, table: str) -> list[str]:
    return [row[1] for row in conn.execute(f"PRAGMA table_info({qtext(table)})").fetchall() if row[1] != "_source_file"]


def aggregate_dict(conn: duckdb.DuckDBPyConnection, table: str, expressions: list[tuple[str, str]]) -> dict[str, Any]:
    sql = ", ".join(f"{expr} AS {qident(alias)}" for alias, expr in expressions)
    row = conn.execute(f"SELECT {sql} FROM {qident(table)}").fetchone()
    return dict(zip((alias for alias, _ in expressions), row))


def date_expression(column: str) -> str:
    c = f"NULLIF(trim({qident(column)}), '')"
    return (
        "coalesce(try_cast(" + c + " AS TIMESTAMP), "
        f"try_strptime({c}, '%d-%m-%Y'), try_strptime({c}, '%d/%m/%Y'), "
        f"try_strptime({c}, '%Y-%m-%d'), try_strptime({c}, '%Y-%m-%d %H:%M:%S'))"
    )


def declared_dictionary_lookup(dictionary_rows: list[dict[str, Any]]) -> dict[tuple[str, str], dict[str, Any]]:
    return {(row["dataset"], row["field"]): row for row in dictionary_rows}


def profile_dataset(
    conn: duckdb.DuckDBPyConnection,
    dataset: str,
    table: str,
    declared_keys: list[str],
    dictionary_lookup: dict[tuple[str, str], dict[str, Any]],
) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    columns = table_columns(conn, table)
    row_count = conn.execute(f"SELECT count(*) FROM {qident(table)}").fetchone()[0]
    file_counts = [
        {"relative_file": Path(row[0]).name, "rows": row[1]}
        for row in conn.execute(f"SELECT _source_file, count(*) FROM {qident(table)} GROUP BY 1 ORDER BY 1").fetchall()
    ]

    null_exprs = []
    approx_exprs = []
    length_exprs = []
    for col in columns:
        c = qident(col)
        null_exprs.append((col, f"count(*) FILTER (WHERE {c} IS NULL OR trim({c}) = '')"))
        approx_exprs.append((col, f"approx_count_distinct(NULLIF(trim({c}), ''))"))
        length_exprs.extend([(f"{col}__min", f"min(length({c}))"), (f"{col}__max", f"max(length({c}))")])
    nulls = aggregate_dict(conn, table, null_exprs)
    approx = aggregate_dict(conn, table, approx_exprs)
    lengths = aggregate_dict(conn, table, length_exprs)

    column_rows: list[dict[str, Any]] = []
    likely_candidates: set[str] = set(declared_keys)
    for ordinal, col in enumerate(columns, start=1):
        non_null = row_count - nulls[col]
        ratio = (approx[col] / non_null) if non_null else 0
        if nulls[col] == 0 and (IDENTIFIER_RE.search(col) or ratio >= 0.90):
            likely_candidates.add(col)
        dictionary = dictionary_lookup.get((dataset, col))
        column_rows.append({
            "dataset": dataset,
            "table": table,
            "ordinal": ordinal,
            "field": col,
            "rows": row_count,
            "null_rows": nulls[col],
            "null_percent": round((nulls[col] / row_count * 100) if row_count else 0, 6),
            "approx_distinct": approx[col],
            "min_length": lengths[f"{col}__min"],
            "max_length": lengths[f"{col}__max"],
            "dictionary_present": dictionary is not None,
            "declared_data_type": dictionary["declared_data_type"] if dictionary else None,
            "declared_primary_key": bool(dictionary and dictionary["declared_primary_key"]),
        })

    candidate_rows: list[dict[str, Any]] = []
    for col in sorted(c for c in likely_candidates if c in columns):
        exact = conn.execute(
            f"SELECT count(DISTINCT NULLIF(trim({qident(col)}), '')) FROM {qident(table)}"
        ).fetchone()[0]
        is_unique = nulls[col] == 0 and exact == row_count
        candidate_rows.append({
            "dataset": dataset,
            "key_fields": col,
            "key_width": 1,
            "declared_by_dld": col in declared_keys,
            "non_null": nulls[col] == 0,
            "distinct_keys": exact,
            "rows": row_count,
            "duplicate_rows": row_count - exact,
            "is_candidate_key": is_unique,
        })

    valid_declared = [col for col in declared_keys if col in columns]
    if len(valid_declared) > 1:
        null_predicate = " OR ".join(f"{qident(col)} IS NULL OR trim({qident(col)}) = ''" for col in valid_declared)
        hash_expr = "hash(" + ",".join(f"NULLIF(trim({qident(col)}), '')" for col in valid_declared) + ")"
        null_count, distinct_count = conn.execute(
            f"SELECT count(*) FILTER (WHERE {null_predicate}), count(DISTINCT {hash_expr}) FROM {qident(table)}"
        ).fetchone()
        candidate_rows.append({
            "dataset": dataset,
            "key_fields": "+".join(valid_declared),
            "key_width": len(valid_declared),
            "declared_by_dld": True,
            "non_null": null_count == 0,
            "distinct_keys": distinct_count,
            "rows": row_count,
            "duplicate_rows": row_count - distinct_count,
            "is_candidate_key": null_count == 0 and distinct_count == row_count,
        })

    row_hash = "hash(" + ",".join(qident(col) for col in columns) + ")"
    distinct_rows = conn.execute(f"SELECT count(DISTINCT {row_hash}) FROM {qident(table)}").fetchone()[0]

    date_rows: list[dict[str, Any]] = []
    for col in columns:
        if not DATE_NAME_RE.search(col):
            continue
        lower_col = col.lower()
        if lower_col.endswith("_year") or lower_col == "year":
            parsed = f"try_cast(NULLIF(trim({qident(col)}), '') AS BIGINT)"
            kind = "year"
        elif lower_col.endswith("_month") or lower_col == "month":
            parsed = f"try_cast(NULLIF(trim({qident(col)}), '') AS BIGINT)"
            kind = "month"
        else:
            parsed = date_expression(col)
            kind = "datetime"
        if kind == "datetime":
            nonblank, parsed_count, minimum, maximum, before_1900, after_2035 = conn.execute(
                f"SELECT count(*) FILTER (WHERE NULLIF(trim({qident(col)}), '') IS NOT NULL), "
                f"count({parsed}), min({parsed}), max({parsed}), "
                f"count(*) FILTER (WHERE {parsed} < TIMESTAMP '1900-01-01'), "
                f"count(*) FILTER (WHERE {parsed} >= TIMESTAMP '2036-01-01') FROM {qident(table)}"
            ).fetchone()
        else:
            nonblank, parsed_count, minimum, maximum = conn.execute(
                f"SELECT count(*) FILTER (WHERE NULLIF(trim({qident(col)}), '') IS NOT NULL), "
                f"count({parsed}), min({parsed}), max({parsed}) FROM {qident(table)}"
            ).fetchone()
            before_1900 = after_2035 = None
        date_rows.append({
            "dataset": dataset,
            "field": col,
            "profile_kind": kind,
            "nonblank_rows": nonblank,
            "parsed_rows": parsed_count,
            "unparsed_rows": nonblank - parsed_count,
            "minimum": minimum.isoformat() if hasattr(minimum, "isoformat") else (str(minimum) if minimum is not None else None),
            "maximum": maximum.isoformat() if hasattr(maximum, "isoformat") else (str(maximum) if maximum is not None else None),
            "before_1900_rows": before_1900,
            "from_2036_rows": after_2035,
        })

    dataset_row = {
        "dataset": dataset,
        "table": table,
        "source_csv_files": len(file_counts),
        "row_count": row_count,
        "column_count": len(columns),
        "distinct_full_row_hashes": distinct_rows,
        "duplicate_full_rows": row_count - distinct_rows,
        "declared_key_fields": declared_keys,
        "csv_files": file_counts,
        "dictionary_only_fields": sorted(
            field for (dict_dataset, field) in dictionary_lookup if dict_dataset == dataset and field not in columns
        ),
        "csv_only_fields": sorted(col for col in columns if (dataset, col) not in dictionary_lookup),
    }
    return dataset_row, column_rows, candidate_rows, date_rows


def norm_expr(column: str) -> str:
    # Excel-backed lookup exports commonly render integral identifiers as
    # strings such as "479.00", while fact exports use "479".
    return f"upper(regexp_replace(trim({qident(column)}), '\\.0+$', ''))"


def test_relationship(conn: duckdb.DuckDBPyConnection, spec: tuple[str, str, str, str, str]) -> dict[str, Any]:
    name, child_table, child_col, parent_table, parent_col = spec
    child_norm = norm_expr(child_col)
    parent_norm = norm_expr(parent_col)
    sql = f"""
        WITH parent_keys AS (
            SELECT DISTINCT {parent_norm} AS key
            FROM {qident(parent_table)}
            WHERE NULLIF(trim({qident(parent_col)}), '') IS NOT NULL
        ), child_keys AS (
            SELECT {child_norm} AS key
            FROM {qident(child_table)}
            WHERE NULLIF(trim({qident(child_col)}), '') IS NOT NULL
        ), joined AS (
            SELECT c.key, p.key IS NOT NULL AS matched
            FROM child_keys c LEFT JOIN parent_keys p USING (key)
        )
        SELECT count(*), count(DISTINCT key),
               count(*) FILTER (WHERE matched),
               count(DISTINCT key) FILTER (WHERE matched),
               count(*) FILTER (WHERE NOT matched),
               count(DISTINCT key) FILTER (WHERE NOT matched)
        FROM joined
    """
    child_rows, child_keys, matched_rows, matched_keys, orphan_rows, orphan_keys = conn.execute(sql).fetchone()
    child_total_rows = conn.execute(f"SELECT count(*) FROM {qident(child_table)}").fetchone()[0]
    return {
        "relationship": name,
        "child_table": child_table,
        "child_field": child_col,
        "parent_table": parent_table,
        "parent_field": parent_col,
        "child_total_rows": child_total_rows,
        "child_null_rows": child_total_rows - child_rows,
        "child_non_null_rows": child_rows,
        "child_distinct_keys": child_keys,
        "matched_rows": matched_rows,
        "matched_distinct_keys": matched_keys,
        "orphan_rows": orphan_rows,
        "orphan_distinct_keys": orphan_keys,
        "row_match_percent": round((matched_rows / child_rows * 100) if child_rows else 100, 6),
        "key_match_percent": round((matched_keys / child_keys * 100) if child_keys else 100, 6),
        "status": "pass" if orphan_rows == 0 else "partial",
    }


def rent_identity(conn: duckdb.DuckDBPyConnection) -> dict[str, Any]:
    table = qident("rent_contracts")
    metrics = conn.execute(f"""
        SELECT
          count(*) AS rows,
          count(*) FILTER (WHERE NULLIF(trim(contract_id), '') IS NULL) AS null_contract_ids,
          count(*) FILTER (WHERE NULLIF(trim(line_number), '') IS NULL) AS null_line_numbers,
          count(DISTINCT NULLIF(trim(contract_id), '')) AS distinct_contract_ids,
          count(DISTINCT hash(NULLIF(trim(contract_id), ''), NULLIF(trim(line_number), ''))) AS distinct_contract_lines
        FROM {table}
    """).fetchone()
    rows, null_contracts, null_lines, contracts, contract_lines = metrics
    multi_line_contracts = conn.execute(f"""
        SELECT count(*) FROM (
          SELECT contract_id FROM {table}
          WHERE NULLIF(trim(contract_id), '') IS NOT NULL
          GROUP BY contract_id HAVING count(DISTINCT line_number) > 1
        )
    """).fetchone()[0]
    cross_file_pairs = conn.execute(f"""
        SELECT count(*) FROM (
          SELECT contract_id, line_number FROM {table}
          GROUP BY contract_id, line_number HAVING count(DISTINCT _source_file) > 1
        )
    """).fetchone()[0]
    no_of_prop_mismatches = conn.execute(f"""
        SELECT count(*) FROM (
          SELECT contract_id,
                 count(DISTINCT line_number) AS observed_lines,
                 max(try_cast(no_of_prop AS BIGINT)) AS declared_properties
          FROM {table} GROUP BY contract_id
        ) WHERE declared_properties IS NOT NULL AND observed_lines <> declared_properties
    """).fetchone()[0]
    composite_valid = null_contracts == 0 and null_lines == 0 and contract_lines == rows
    return {
        "rows": rows,
        "null_contract_ids": null_contracts,
        "null_line_numbers": null_lines,
        "distinct_contract_ids": contracts,
        "rows_beyond_contract_id_uniqueness": rows - contracts,
        "distinct_contract_id_line_number_pairs": contract_lines,
        "duplicate_contract_id_line_number_rows": rows - contract_lines,
        "contracts_with_multiple_line_numbers": multi_line_contracts,
        "contract_line_pairs_crossing_source_files": cross_file_pairs,
        "contracts_where_no_of_prop_differs_from_observed_lines": no_of_prop_mismatches,
        "dictionary_declared_key": ["contract_id", "line_number"],
        "contract_identity": ["contract_id"],
        "rent_property_row_identity": ["contract_id", "line_number"],
        "composite_key_valid": composite_valid,
        "finding": (
            "contract_id identifies an Ejari contract header; (contract_id, line_number) identifies each property line. "
            "A contract_id-only upsert would collapse multi-property contracts."
        ),
    }


def classify_field(dataset: str, field: str) -> tuple[str, str]:
    lower = field.lower()
    if lower in INTERNAL_EXACT or lower in INTERNAL_CONTACT:
        return "internal", "Operational identifier/contact or ingestion metadata; retain in the controlled data layer."
    if lower == "load_timestamp" or lower.endswith("_property_id"):
        return "internal", "Source-system lineage or property identifier; not a public product field."
    if lower in {"land_number", "land_sub_number", "unit_number", "building_number", "munc_number", "munc_zip_code"}:
        return "internal", "Granular property locator; not suitable for public row-level exposure."
    if dataset == "Real Estate Brokers" and lower == "gender":
        return "aggregate-only", "Person-level demographic attribute; publish only as a sufficiently large aggregate."
    if dataset in GRANULAR_PROPERTY_DATASETS:
        if lower.endswith(("_name_en", "_name_ar", "_type_en", "_type_ar", "_usage_en", "_usage_ar")) or lower in {
            "area_id", "property_type_id", "property_sub_type_id", "reg_type_id", "is_free_hold", "is_lease_hold",
            "contract_reg_type_id", "tenant_type_id", "ejari_property_type_id", "ejari_property_sub_type_id",
            "ejari_bus_property_type_id", "rooms", "rooms_en", "rooms_ar", "has_parking", "is_registered",
        }:
            return "public", "Non-personal market dimension suitable for labels, filters, and methodology."
        return "aggregate-only", "Granular property/transaction/rent measure; publish only through controlled aggregates."
    if dataset == "Property Map Requests" and lower in {"request_date", "no_of_siteplans"}:
        return "aggregate-only", "Operational service activity; publish only as aggregated demand statistics."
    if any(token in lower for token in ("amount", "worth", "price", "service_cost")) and dataset not in {
        "Residential Sale Index", "Owners Association Service Charges"
    }:
        return "aggregate-only", "Financial measure; publish as an aggregate unless explicitly approved."
    return "public", "Open, non-personal registry/dimension field suitable for public use with DLD attribution."


def build_classifications(dataset_columns: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for column in dataset_columns:
        classification, rationale = classify_field(column["dataset"], column["field"])
        rows.append({
            "dataset": column["dataset"],
            "field": column["field"],
            "classification": classification,
            "rationale": rationale,
            "source_dictionary_classification": "Open" if column["dictionary_present"] else None,
        })
    return rows


def markdown_report(
    generated_at: str,
    inventory: list[dict[str, Any]],
    datasets: list[dict[str, Any]],
    candidates: list[dict[str, Any]],
    dates: list[dict[str, Any]],
    relationships: list[dict[str, Any]],
    rent: dict[str, Any],
    classifications: list[dict[str, Any]],
) -> str:
    total_rows = sum(row["row_count"] for row in datasets)
    duplicate_rows = sum(row["duplicate_full_rows"] for row in datasets)
    relation_passes = sum(row["status"] == "pass" for row in relationships)
    classes = Counter(row["classification"] for row in classifications)
    largest = sorted(datasets, key=lambda row: row["row_count"], reverse=True)[:8]
    date_flags = [
        row for row in dates
        if row["profile_kind"] == "datetime" and ((row["before_1900_rows"] or 0) + (row["from_2036_rows"] or 0) > 0)
    ]
    valid_keys = [row for row in candidates if row["is_candidate_key"]]
    lines = [
        "# DLD Phase 0 inventory and audit",
        "",
        f"Generated: `{generated_at}`",
        "",
        "Phase 0 is a local, read-only assessment. No Supabase or production writes are performed.",
        "Committed outputs contain metadata and counts only; the source CSV/XLSX rows stay outside Git.",
        "",
        "## Scope and headline findings",
        "",
        f"- Inventoried **{len(inventory)} files**: **{sum(r['extension'] == '.csv' for r in inventory)} CSV** exports and **{sum(r['extension'] == '.xlsx' for r in inventory)} Excel** dictionaries.",
        f"- Read **{len(datasets)} datasets** containing **{total_rows:,} source rows**.",
        f"- Detected **{duplicate_rows:,} exact duplicate rows** by full-row hash across all datasets.",
        f"- Confirmed **{len(valid_keys)} candidate-key tests** as unique and non-null; all tested key results are in the machine-readable report.",
        f"- Relationship checks: **{relation_passes}/{len(relationships)}** have complete non-null key coverage; partial checks retain quantitative orphan rates.",
        f"- Rent identity: `contract_id` is the contract header; **(`contract_id`, `line_number`)** is the rent-property row identity. Composite-key validity: **{rent['composite_key_valid']}**.",
        "- The existing `contract_id`-only rent ingestion design would collapse property lines for multi-property contracts and must not be used for a production load without redesign.",
        "",
        "## Largest datasets",
        "",
        "| Dataset | Rows | Columns | CSV chunks | Exact duplicate rows |",
        "| --- | ---: | ---: | ---: | ---: |",
    ]
    lines.extend(
        f"| {row['dataset']} | {row['row_count']:,} | {row['column_count']} | {row['source_csv_files']} | {row['duplicate_full_rows']:,} |"
        for row in largest
    )
    lines.extend([
        "",
        "## Rent-contract identity",
        "",
        f"- Rows: **{rent['rows']:,}**; distinct contracts: **{rent['distinct_contract_ids']:,}**.",
        f"- Contracts with multiple property lines: **{rent['contracts_with_multiple_line_numbers']:,}**.",
        f"- Duplicate (`contract_id`, `line_number`) rows: **{rent['duplicate_contract_id_line_number_rows']:,}**.",
        f"- Contract-line pairs crossing CSV chunk boundaries: **{rent['contract_line_pairs_crossing_source_files']:,}**.",
        f"- Contracts whose declared `no_of_prop` differs from observed lines: **{rent['contracts_where_no_of_prop_differs_from_observed_lines']:,}**.",
        "",
        "## Relationship coverage",
        "",
        "| Relationship | Non-null child rows | Null child rows | Row match | Orphan keys | Status |",
        "| --- | ---: | ---: | ---: | ---: | --- |",
    ])
    lines.extend(
        f"| `{row['relationship']}` | {row['child_non_null_rows']:,} | {row['child_null_rows']:,} | {row['row_match_percent']:.3f}% | {row['orphan_distinct_keys']:,} | {row['status']} |"
        for row in relationships
    )
    lines.extend([
        "",
        "## Date-quality flags",
        "",
        "Dates before 1900 or from 2036 onward are retained as source values but flagged for review; the threshold is a quality screen, not a deletion rule.",
        "",
        "| Dataset.field | Range | Before 1900 | From 2036 |",
        "| --- | --- | ---: | ---: |",
    ])
    lines.extend(
        f"| `{row['dataset']}.{row['field']}` | {row['minimum']} to {row['maximum']} | {row['before_1900_rows']:,} | {row['from_2036_rows']:,} |"
        for row in date_flags
    )
    if not date_flags:
        lines.append("| — | No threshold exceptions | 0 | 0 |")
    lines.extend([
        "",
        "## Publication classification",
        "",
        f"- Public: **{classes['public']}** fields.",
        f"- Aggregate-only: **{classes['aggregate-only']}** fields.",
        f"- Internal: **{classes['internal']}** fields.",
        "",
        "DLD's dictionaries label the exports `Open`; this audit adds a stricter product-exposure classification.",
        "Open-data status does not mean every granular identifier, contact field, or property-level measure should be emitted by the website.",
        "",
        "## Machine-readable evidence",
        "",
        "See `reports/dld/phase0/` for the file inventory and hashes, normalized dictionaries, dataset and column profiles, date ranges, candidate keys, relationship tests, rent identity, and field classifications.",
        "The ignored local DuckDB database is a reproducible working product, not a deployable artifact.",
        "",
        "## Phase boundary",
        "",
        "No raw source files were copied into the repository, no Supabase code or data was changed, and nothing was committed, pushed, deployed, or published.",
        "",
    ])
    return "\n".join(lines)


def classification_markdown(classifications: list[dict[str, Any]]) -> str:
    counts = Counter(row["classification"] for row in classifications)
    return "\n".join([
        "# DLD field publication classification",
        "",
        "This is a product-exposure policy layered on top of the source dictionaries' `Open` label.",
        "",
        "- **Public**: safe for direct labels, registry facts, dimensions, and already-aggregated statistics, with DLD attribution.",
        "- **Aggregate-only**: usable for statistics after grouping and disclosure controls; do not expose source rows.",
        "- **Internal**: join, lineage, contact, or granular locator fields retained only in controlled processing.",
        "",
        f"Current totals: public **{counts['public']}**, aggregate-only **{counts['aggregate-only']}**, internal **{counts['internal']}**.",
        "",
        "The complete per-field decision and rationale is in `reports/dld/phase0/field_classification.csv`.",
        "",
    ])


def dictionary_markdown(datasets: list[dict[str, Any]], dictionary_rows: list[dict[str, Any]]) -> str:
    lines = [
        "# DLD attribute dictionaries",
        "",
        f"All **{len(set(row['workbook'] for row in dictionary_rows))}** Excel dictionaries were read in full and normalized into `reports/dld/phase0/attribute_dictionary.csv`.",
        "",
        "| Dataset | Dictionary fields | CSV columns | Dictionary-only | CSV-only | Declared key |",
        "| --- | ---: | ---: | --- | --- | --- |",
    ]
    counts = Counter(row["dataset"] for row in dictionary_rows)
    for dataset in datasets:
        lines.append(
            f"| {dataset['dataset']} | {counts[dataset['dataset']]} | {dataset['column_count']} | "
            f"{', '.join(dataset['dictionary_only_fields']) or '—'} | {', '.join(dataset['csv_only_fields']) or '—'} | "
            f"{', '.join(dataset['declared_key_fields']) or 'none declared'} |"
        )
    lines.extend(["", "`load_timestamp` is expected to be CSV-only because it is ingestion lineage rather than a DLD business attribute.", ""])
    return "\n".join(lines)


def report_checksums(report_dir: Path) -> list[dict[str, Any]]:
    rows = []
    for path in sorted(report_dir.glob("*")):
        if path.is_file() and path.name not in {"report_manifest.json", "verification.json"}:
            rows.append({"file": path.name, "bytes": path.stat().st_size, "sha256": sha256_file(path)})
    return rows


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, required=True, help="Read-only DLD source directory")
    parser.add_argument("--reports", type=Path, default=Path("reports/dld/phase0"))
    parser.add_argument("--database", type=Path, default=Path("data/dld/local/phase0.duckdb"))
    parser.add_argument("--reuse-db", action="store_true", help="Reuse already materialized DuckDB tables")
    args = parser.parse_args()

    source = args.source.resolve()
    if not source.is_dir():
        parser.error(f"Source directory does not exist: {source}")
    initial_stats = {p: (p.stat().st_size, p.stat().st_mtime_ns) for p in source_files(source)}
    generated_at = utc_now()
    started = time.monotonic()

    print("Inventorying and hashing source files...", flush=True)
    inventory = inventory_source(source)
    if len(inventory) != 62 or sum(row["extension"] == ".xlsx" for row in inventory) != 25:
        raise ValueError("Expected 62 files including 25 Excel dictionaries")

    print("Reading all Excel dictionaries...", flush=True)
    dictionary_rows, declared_keys = read_dictionaries(source)
    dictionary_lookup = declared_dictionary_lookup(dictionary_rows)
    csv_groups = group_csvs(source)
    if set(csv_groups) != set(DATASET_TABLES):
        raise ValueError(f"Dataset mismatch: source={sorted(csv_groups)}, configured={sorted(DATASET_TABLES)}")

    conn = connect_database(args.database)
    datasets: list[dict[str, Any]] = []
    columns: list[dict[str, Any]] = []
    candidates: list[dict[str, Any]] = []
    dates: list[dict[str, Any]] = []
    for index, dataset in enumerate(DATASET_TABLES, start=1):
        table = DATASET_TABLES[dataset]
        print(f"[{index:02d}/25] materialize/profile {dataset}", flush=True)
        materialize(conn, table, csv_groups[dataset], args.reuse_db)
        dataset_row, column_rows, candidate_rows, date_rows = profile_dataset(
            conn, dataset, table, declared_keys.get(dataset, []), dictionary_lookup
        )
        datasets.append(dataset_row)
        columns.extend(column_rows)
        candidates.extend(candidate_rows)
        dates.extend(date_rows)

    print("Testing cross-dataset relationships...", flush=True)
    relationships = []
    for spec in RELATIONSHIPS:
        print(f"  {spec[0]}", flush=True)
        relationships.append(test_relationship(conn, spec))

    print("Testing rent-contract identity...", flush=True)
    rent = rent_identity(conn)
    classifications = build_classifications(columns)
    conn.close()

    current_files = source_files(source)
    final_stats = {p: (p.stat().st_size, p.stat().st_mtime_ns) for p in current_files}
    source_unchanged = initial_stats == final_stats
    if not source_unchanged:
        raise RuntimeError("Source folder changed during the audit; reports were not written")

    report_dir = args.reports
    write_json(report_dir / "inventory.json", inventory)
    write_csv(report_dir / "inventory.csv", inventory)
    write_json(report_dir / "attribute_dictionary.json", dictionary_rows)
    write_csv(report_dir / "attribute_dictionary.csv", dictionary_rows)
    write_json(report_dir / "datasets.json", datasets)
    write_csv(report_dir / "datasets.csv", datasets, [
        "dataset", "table", "source_csv_files", "row_count", "column_count",
        "distinct_full_row_hashes", "duplicate_full_rows",
    ])
    write_json(report_dir / "columns.json", columns)
    write_csv(report_dir / "columns.csv", columns)
    write_json(report_dir / "candidate_keys.json", candidates)
    write_csv(report_dir / "candidate_keys.csv", candidates)
    write_json(report_dir / "date_ranges.json", dates)
    write_csv(report_dir / "date_ranges.csv", dates)
    write_json(report_dir / "relationships.json", relationships)
    write_csv(report_dir / "relationships.csv", relationships)
    write_json(report_dir / "rent_identity.json", rent)
    write_json(report_dir / "field_classification.json", classifications)
    write_csv(report_dir / "field_classification.csv", classifications)

    docs_dir = Path("docs/dld")
    docs_dir.mkdir(parents=True, exist_ok=True)
    (docs_dir / "PHASE-0-AUDIT.md").write_text(
        markdown_report(generated_at, inventory, datasets, candidates, dates, relationships, rent, classifications), encoding="utf-8"
    )
    (docs_dir / "FIELD-CLASSIFICATION.md").write_text(classification_markdown(classifications), encoding="utf-8")
    (docs_dir / "ATTRIBUTE-DICTIONARIES.md").write_text(dictionary_markdown(datasets, dictionary_rows), encoding="utf-8")

    metadata = {
        "phase": 0,
        "generated_at": generated_at,
        "duration_seconds": round(time.monotonic() - started, 3),
        "source_directory_name": source.name,
        "source_unchanged_during_run": source_unchanged,
        "file_count": len(inventory),
        "csv_count": sum(row["extension"] == ".csv" for row in inventory),
        "xlsx_count": sum(row["extension"] == ".xlsx" for row in inventory),
        "dataset_count": len(datasets),
        "total_rows": sum(row["row_count"] for row in datasets),
        "supabase_writes": 0,
        "raw_rows_in_reports": 0,
        "duckdb_version": duckdb.__version__,
    }
    write_json(report_dir / "audit_metadata.json", metadata)
    write_json(report_dir / "report_manifest.json", report_checksums(report_dir))
    print(f"Phase 0 reports written to {report_dir} in {metadata['duration_seconds']:.1f}s", flush=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
