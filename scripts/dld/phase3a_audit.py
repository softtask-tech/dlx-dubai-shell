#!/usr/bin/env python3
"""Build the local, aggregate-only DLD Phase 3A market-data audit.

The script reuses the ignored Phase 0 DuckDB database and never emits source
rows. Versioned outputs contain counts, schemas, policies and aggregate
distributions only. The source export and database are opened read-only.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import zipfile
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

import duckdb
from openpyxl import load_workbook

from phase0_audit import DATASET_TABLES


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_SOURCE = Path(r"C:\Users\adams\Downloads\DLD DATA COD")
DEFAULT_DB = ROOT / "data" / "dld" / "local" / "phase0.duckdb"
DEFAULT_REPORTS = ROOT / "reports" / "dld" / "phase3a"
PHASE0_REPORTS = ROOT / "reports" / "dld" / "phase0"
METHODOLOGY_VERSION = "dld-market-audit-phase3a-v1"
EXPECTED_DATASETS = 25
EXPECTED_FILES = 62
SQM_TO_SQFT = 10.7639104167
MARKET_SALE_PROCEDURE_IDS = {11, 41, 102, 460}


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True, default=json_default) + "\n",
        encoding="utf-8",
    )


def json_default(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    raise TypeError(f"Unsupported JSON value: {type(value).__name__}")


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fields is None:
        fields = list(rows[0]) if rows else []
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(8 * 1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def scalar(conn: duckdb.DuckDBPyConnection, sql: str) -> Any:
    return conn.execute(sql).fetchone()[0]


def row_dict(conn: duckdb.DuckDBPyConnection, sql: str) -> dict[str, Any]:
    cursor = conn.execute(sql)
    row = cursor.fetchone()
    return dict(zip([item[0] for item in cursor.description], row))


def rows_dict(conn: duckdb.DuckDBPyConnection, sql: str) -> list[dict[str, Any]]:
    cursor = conn.execute(sql)
    fields = [item[0] for item in cursor.description]
    return [dict(zip(fields, row)) for row in cursor.fetchall()]


def date_expr(field: str) -> str:
    source = f'nullif(trim("{field}"), \'\')'
    return (
        f"coalesce(try_cast({source} as timestamp), try_strptime({source}, '%d-%m-%Y'), "
        f"try_strptime({source}, '%d/%m/%Y'), try_strptime({source}, '%Y-%m-%d'), "
        f"try_strptime({source}, '%Y-%m-%d %H:%M:%S'))"
    )


def number_expr(field: str) -> str:
    return f'try_cast(nullif(trim("{field}"), \'\') as double)'


def detect_encoding(path: Path) -> str:
    sample = path.read_bytes()[: 128 * 1024]
    if sample.startswith(b"\xef\xbb\xbf"):
        return "UTF-8 with BOM"
    try:
        sample.decode("utf-8")
        return "UTF-8"
    except UnicodeDecodeError:
        try:
            sample.decode("cp1256")
            return "Windows-1256 (sample inference)"
        except UnicodeDecodeError:
            return "Unknown/needs review"


def export_date_from_name(name: str) -> str | None:
    matches = re.findall(r"20\d{2}-\d{2}-\d{2}", name)
    return max(matches) if matches else None


def source_inventory(source: Path, phase0_datasets: list[dict[str, Any]]) -> list[dict[str, Any]]:
    profiles = {row["dataset"]: row for row in phase0_datasets}
    result: list[dict[str, Any]] = []
    for folder in sorted(path for path in source.iterdir() if path.is_dir()):
        files = []
        export_dates: list[str] = []
        formats: Counter[str] = Counter()
        csv_encodings: set[str] = set()
        sheet_names: set[str] = set()
        compressed = 0
        uncompressed = 0
        for path in sorted(item for item in folder.iterdir() if item.is_file()):
            relative = path.relative_to(source).as_posix()
            suffix = path.suffix.lower().lstrip(".")
            formats[suffix] += 1
            size = path.stat().st_size
            compressed += size
            file_uncompressed = size
            encoding = "binary OOXML" if suffix == "xlsx" else "binary Excel" if suffix == "xls" else detect_encoding(path)
            sheets: list[dict[str, Any]] = []
            if suffix == "csv":
                csv_encodings.add(encoding)
            elif suffix == "xlsx":
                with zipfile.ZipFile(path) as archive:
                    file_uncompressed = sum(item.file_size for item in archive.infolist())
                workbook = load_workbook(path, read_only=True, data_only=True)
                for sheet in workbook.worksheets:
                    sheets.append({"name": sheet.title, "rows_including_header": sheet.max_row, "columns": sheet.max_column})
                    sheet_names.add(sheet.title)
                workbook.close()
            uncompressed += file_uncompressed
            export_date = export_date_from_name(path.name)
            if export_date:
                export_dates.append(export_date)
            files.append(
                {
                    "path": relative,
                    "format": suffix.upper(),
                    "compressed_bytes": size,
                    "uncompressed_bytes": file_uncompressed,
                    "encoding": encoding,
                    "sheets": sheets,
                    "sha256": sha256_file(path),
                    "source_export_date": export_date,
                }
            )
        profile = profiles[folder.name]
        result.append(
            {
                "dataset": folder.name,
                "table": profile["table"],
                "dataset_kind": dataset_kind(folder.name),
                "file_count": len(files),
                "formats": dict(sorted(formats.items())),
                "compressed_bytes": compressed,
                "uncompressed_bytes": uncompressed,
                "encodings": sorted(csv_encodings) or ["binary OOXML"],
                "sheet_names": sorted(sheet_names),
                "row_count": profile["row_count"],
                "column_count": profile["column_count"],
                "candidate_primary_keys": [],
                "source_export_date": max(export_dates) if export_dates else None,
                "duplicate_decode_risks": [
                    "CSV chunks are one logical snapshot; do not append a repeated export.",
                    "Identifiers are strings in the raw landing layer; normalize before comparisons.",
                ],
                "files": files,
            }
        )
    return result


def dataset_kind(name: str) -> str:
    if name.startswith("Lookup "):
        return "lookup"
    if name in {"Real Estate Transactions", "Property Valuation Records", "Property Map Requests"}:
        return "event"
    if name == "Rent Contracts":
        return "contract-line snapshot"
    if name == "Residential Sale Index":
        return "aggregate time-series snapshot"
    if name in {"Land Registry", "Registered Freehold Real Estate Units", "Building and Property Project Records"}:
        return "property/registry snapshot"
    return "directory/snapshot"


def add_candidate_keys(inventory: list[dict[str, Any]], keys: list[dict[str, Any]]) -> None:
    by_dataset: dict[str, list[str]] = {}
    for row in keys:
        if row["is_candidate_key"]:
            by_dataset.setdefault(row["dataset"], []).append(row["key_fields"])
    for dataset in inventory:
        dataset["candidate_primary_keys"] = sorted(by_dataset.get(dataset["dataset"], []))


def add_date_coverage(inventory: list[dict[str, Any]], date_ranges: list[dict[str, Any]]) -> None:
    by_dataset: dict[str, list[dict[str, Any]]] = {}
    for row in date_ranges:
        by_dataset.setdefault(row["dataset"], []).append(row)
    for dataset in inventory:
        dataset["date_coverage"] = sorted(
            by_dataset.get(dataset["dataset"], []), key=lambda row: row["field"]
        )


def schema_fingerprints(conn: duckdb.DuckDBPyConnection, inventory: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for dataset in inventory:
        columns = rows_dict(conn, f'DESCRIBE "{dataset["table"]}"')
        signature = [{"name": row["column_name"], "type": row["column_type"]} for row in columns]
        digest = hashlib.sha256(json.dumps(signature, separators=(",", ":")).encode()).hexdigest()
        result.append({"dataset": dataset["dataset"], "table": dataset["table"], "sha256": digest, "columns": signature})
    return result


def market_sale_sql(alias: str = "") -> str:
    prefix = f"{alias}." if alias else ""
    ids = ", ".join(str(value) for value in sorted(MARKET_SALE_PROCEDURE_IDS))
    return (
        f"try_cast({prefix}trans_group_id as integer) = 1 and "
        f"try_cast({prefix}procedure_id as integer) in ({ids})"
    )


def procedure_policy(conn: duckdb.DuckDBPyConnection) -> list[dict[str, Any]]:
    rows = rows_dict(
        conn,
        """
        select try_cast(p.group_id as integer) group_id,
          try_cast(p.procedure_id as integer) procedure_id,
          p.name_en official_name_en, p.name_ar official_name_ar,
          try_cast(p.is_pre_registration as integer) = 1 is_pre_registration,
          count(t.transaction_id) transaction_rows
        from transaction_procedures p
        left join transactions t
          on try_cast(t.trans_group_id as integer) = try_cast(p.group_id as integer)
         and try_cast(t.procedure_id as integer) = try_cast(p.procedure_id as integer)
        group by 1,2,3,4,5 order by 1,2
        """,
    )
    for row in rows:
        group_id = row["group_id"]
        procedure_id = row["procedure_id"]
        if group_id == 1 and procedure_id in MARKET_SALE_PROCEDURE_IDS:
            classification = "market_sale_candidate"
            eligible = True
            rationale = "Conservative exact allowlist of procedures whose official English name directly describes a sale."
        elif group_id == 1 and "lease to own" in (row["official_name_en"] or "").lower():
            classification = "lease_to_own_non_standard"
            eligible = False
            rationale = "Lease-to-own is economically distinct and excluded from sale-market metrics."
        elif group_id == 1:
            classification = "administrative_or_ambiguous_sales_group"
            eligible = False
            rationale = "The official label indicates development/land registration or is otherwise insufficient to prove an arm's-length sale."
        elif group_id == 2:
            classification = "mortgage_or_finance"
            eligible = False
            rationale = "Official transaction group is Mortgages."
        elif group_id == 3:
            classification = "gift_or_grant"
            eligible = False
            rationale = "Official transaction group is Gifts."
        else:
            classification = "unknown_review_required"
            eligible = False
            rationale = "Unknown transaction group; publication must fail pending review."
        row.update(
            {
                "classification": classification,
                "eligible_for_sale_market_metrics": eligible,
                "policy_version": METHODOLOGY_VERSION,
                "rationale": rationale,
            }
        )
    return rows


def transaction_profile(conn: duckdb.DuckDBPyConnection) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    d = date_expr("instance_date")
    worth = number_expr("actual_worth")
    area = number_expr("procedure_area")
    psm = number_expr("meter_sale_price")
    summary = row_dict(
        conn,
        f"""
        select count(*) total_rows,
          count(distinct transaction_id) distinct_transaction_ids,
          count(*) - count(distinct transaction_id) duplicate_transaction_ids,
          count(*) filter (where {d} is null) invalid_dates,
          min({d}) valid_date_min, max({d}) valid_date_max,
          count(*) filter (where {d} < date '1900-01-01') dates_before_1900,
          count(*) filter (where {d} > date '2026-09-04') dates_after_export,
          count(*) filter (where {worth} is null) invalid_or_missing_worth,
          count(*) filter (where {worth} = 0) zero_worth,
          count(*) filter (where {worth} < 0) negative_worth,
          count(*) filter (where {area} is null) invalid_or_missing_area,
          count(*) filter (where {area} = 0) zero_area,
          count(*) filter (where {area} < 0) negative_area,
          count(*) filter (where lower(trim(trans_group_en)) = 'sales') sale_rows,
          count(*) filter (where {market_sale_sql()}) market_sale_candidate_rows,
          count(*) filter (where lower(trim(trans_group_en)) = 'sales' and not ({market_sale_sql()})) non_standard_or_ambiguous_sales_group_rows,
          count(*) filter (where lower(trim(trans_group_en)) = 'mortgages') mortgage_rows,
          count(*) filter (where lower(trim(trans_group_en)) = 'gifts') gift_rows,
          quantile_cont({worth}, 0.01) worth_p01,
          quantile_cont({worth}, 0.5) worth_median,
          quantile_cont({worth}, 0.99) worth_p99,
          quantile_cont({worth}, 0.999) worth_p999,
          max({worth}) worth_max,
          quantile_cont({area}, 0.01) area_sqm_p01,
          quantile_cont({area}, 0.5) area_sqm_median,
          quantile_cont({area}, 0.99) area_sqm_p99,
          quantile_cont({psm}, 0.5) source_price_per_sqm_median,
          count(*) filter (
            where {worth} > 0 and {area} > 0 and {psm} > 0
              and abs(({worth} / {area}) - {psm}) / {psm} <= 0.01
          ) meter_price_reconciles_within_1pct,
          count(*) filter (where {worth} > 0 and {area} > 0 and {psm} > 0) meter_price_comparable_rows
        from transactions
        """,
    )
    summary.update(
        {
            "row_semantics": "One registered transaction event per unique transaction_id.",
            "amount_semantics": "actual_worth is the transaction property value; currency is not explicit in the export dictionary and needs authority confirmation before publication.",
            "area_semantics": "procedure_area is explicitly documented as square metres; source meter_sale_price is AED-per-square-metre only if currency is confirmed.",
            "sale_definition": "Only the versioned procedure-policy allowlist within transaction group 1 is eligible for market-sale metrics. Other Sales-group procedures remain excluded pending semantic confirmation.",
            "primary_secondary_market": "unsupported: reg_type distinguishes existing from off-plan, not seller identity or primary/secondary market.",
            "developer_presence": "not present directly; may be inherited only through a unique, matched project_number relationship.",
            "participant_data": "only aggregate party counts are supplied; no buyer/seller names are present in this export.",
        }
    )
    periods = rows_dict(
        conn,
        f"""
        select strftime({d}, '%Y-%m') period,
          count(*) records,
          count(*) filter (where lower(trim(trans_group_en)) = 'sales') sales,
          count(*) filter (where {market_sale_sql()}) market_sale_candidates,
          count(*) filter (where lower(trim(trans_group_en)) = 'mortgages') mortgages,
          count(*) filter (where lower(trim(trans_group_en)) = 'gifts') gifts
        from transactions where {d} between date '1900-01-01' and date '2026-09-04'
        group by 1 order by 1
        """,
    )
    categories: list[dict[str, Any]] = []
    for dimension, field in [
        ("transaction_group", "trans_group_en"),
        ("procedure", "procedure_name_en"),
        ("registration_type", "reg_type_en"),
        ("property_type", "property_type_en"),
        ("property_sub_type", "property_sub_type_en"),
        ("usage", "property_usage_en"),
    ]:
        categories.extend(
            {
                "dimension": dimension,
                "value": row["category_value"],
                "records": row["records"],
            }
            for row in rows_dict(
                conn,
                f'''select coalesce(nullif(trim("{field}"), ''), '[missing]') category_value, count(*) records
                    from transactions group by 1 order by records desc, category_value''',
            )
        )
    return summary, periods, categories


def residential_index_profile(conn: duckdb.DuckDBPyConnection) -> dict[str, Any]:
    numeric_fields = [
        row[0]
        for row in conn.execute("describe residential_sale_index").fetchall()
        if row[0] not in {"first_date_of_month", "load_timestamp", "_source_file"}
    ]
    expressions = []
    for field in numeric_fields:
        value = number_expr(field)
        expressions.extend(
            [
                f'count(*) filter (where nullif(trim("{field}"), \'\') is null) "{field}__missing"',
                f'count(*) filter (where nullif(trim("{field}"), \'\') is not null and {value} is null) "{field}__invalid"',
            ]
        )
    stats = row_dict(conn, f"select {', '.join(expressions)} from residential_sale_index")
    coverage = []
    for field in numeric_fields:
        coverage.append(
            {
                "field": field,
                "missing_rows": stats[f"{field}__missing"],
                "invalid_numeric_rows": stats[f"{field}__invalid"],
            }
        )
    date_value = date_expr("first_date_of_month")
    summary = row_dict(
        conn,
        f"""
        select count(*) total_rows, count(distinct first_date_of_month) distinct_period_keys,
          count(*) filter (where {date_value} is null) invalid_period_keys,
          min({date_value}) period_min, max({date_value}) period_max,
          count(*) filter (where day({date_value}) <> 1) non_month_start_keys
        from residential_sale_index
        """,
    )
    summary.update(
        {
            "row_semantics": "One published residential sale-index month; first_date_of_month is unique.",
            "frequency": "monthly rows with monthly, quarterly and yearly index/value columns supplied together",
            "metric_status": "source aggregate for separate methodology review; not mixed into transaction-derived Phase 3B metrics",
            "coverage_by_value_field": coverage,
        }
    )
    return summary


def rent_profile(conn: duckdb.DuckDBPyConnection) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    start = date_expr("contract_start_date")
    end = date_expr("contract_end_date")
    annual = number_expr("annual_amount")
    amount = number_expr("contract_amount")
    area = number_expr("actual_area")
    conn.execute("drop table if exists phase3a_contract_rollup")
    conn.execute(
        f"""
        create temporary table phase3a_contract_rollup as
        select contract_id, count(*) lines,
          min(nullif(trim(area_id), '')) min_area, max(nullif(trim(area_id), '')) max_area,
          min(nullif(trim(project_number), '')) min_project, max(nullif(trim(project_number), '')) max_project,
          min(nullif(trim(annual_amount), '')) min_annual_raw, max(nullif(trim(annual_amount), '')) max_annual_raw,
          min(nullif(trim(contract_amount), '')) min_amount_raw, max(nullif(trim(contract_amount), '')) max_amount_raw,
          min(nullif(trim(contract_start_date), '')) min_start_raw, max(nullif(trim(contract_start_date), '')) max_start_raw,
          min(nullif(trim(contract_end_date), '')) min_end_raw, max(nullif(trim(contract_end_date), '')) max_end_raw,
          min(nullif(trim(contract_reg_type_id), '')) min_type, max(nullif(trim(contract_reg_type_id), '')) max_type,
          max(try_cast(no_of_prop as bigint)) declared_properties,
          max({annual}) annual_amount,
          max({amount}) contract_amount,
          min({start}) start_date,
          max({end}) end_date
        from rent_contracts group by contract_id
        """
    )
    summary = row_dict(
        conn,
        f"""
        select (select count(*) from rent_contracts) total_property_lines,
          count(*) distinct_contracts,
          (select count(*) - count(distinct contract_id || '|' || line_number) from rent_contracts) duplicate_contract_line_keys,
          count(*) filter (where lines > 1) multi_line_contracts,
          sum(lines) filter (where lines > 1) lines_in_multi_line_contracts,
          count(*) filter (where min_area is distinct from max_area) multi_area_contracts,
          count(*) filter (where min_project is distinct from max_project) multi_project_contracts,
          count(*) filter (where min_annual_raw is distinct from max_annual_raw or min_amount_raw is distinct from max_amount_raw or min_start_raw is distinct from max_start_raw or min_end_raw is distinct from max_end_raw or min_type is distinct from max_type) inconsistent_header_contracts,
          count(*) filter (where declared_properties <> lines) declared_property_count_mismatches,
          count(*) filter (where start_date is null or end_date is null) invalid_contract_dates,
          min(start_date) valid_start_min, max(start_date) valid_start_max,
          min(end_date) valid_end_min, max(end_date) valid_end_max,
          count(*) filter (where start_date > date '2026-09-04') starts_after_export,
          count(*) filter (where end_date > date '2026-09-04') ends_after_export,
          count(*) filter (where start_date >= date '2036-01-01') starts_from_2036,
          count(*) filter (where end_date >= date '2036-01-01') ends_from_2036,
          count(*) filter (where end_date < start_date) negative_duration_contracts,
          count(*) filter (where date_diff('day', start_date, end_date) + 1 between 1 and 3660) plausible_duration_contracts,
          count(*) filter (where date_diff('day', start_date, end_date) + 1 not between 1 and 3660) implausible_duration_contracts,
          quantile_cont(date_diff('day', start_date, end_date) + 1, 0.5) duration_days_median,
          count(*) filter (where annual_amount is null) invalid_or_missing_annual_amount,
          count(*) filter (where annual_amount = 0) zero_annual_amount,
          count(*) filter (where annual_amount < 0) negative_annual_amount,
          quantile_cont(annual_amount, 0.01) annual_amount_p01,
          quantile_cont(annual_amount, 0.5) annual_amount_median,
          quantile_cont(annual_amount, 0.99) annual_amount_p99,
          quantile_cont(annual_amount, 0.999) annual_amount_p999,
          max(annual_amount) annual_amount_max,
          count(*) filter (where try_cast(max_type as bigint) = 1) new_contracts,
          count(*) filter (where try_cast(max_type as bigint) = 2) renewed_contracts,
          count(*) filter (
            where annual_amount > 0 and contract_amount > 0
              and date_diff('day', start_date, end_date) + 1 between 1 and 3660
              and abs(annual_amount - contract_amount * 365.25 / (date_diff('day', start_date, end_date) + 1)) / annual_amount <= 0.02
          ) annualization_reconciles_within_2pct,
          count(*) filter (
            where annual_amount > 0 and contract_amount > 0
              and date_diff('day', start_date, end_date) + 1 between 1 and 3660
          ) annualization_comparable_contracts
        from phase3a_contract_rollup
        """,
    )
    line_quality = row_dict(
        conn,
        f"""
        select count(*) filter (where {area} is null) invalid_or_missing_area,
          count(*) filter (where {area} = 0) zero_area,
          count(*) filter (where {area} < 0) negative_area,
          quantile_cont({area}, 0.01) area_p01,
          quantile_cont({area}, 0.5) area_median,
          quantile_cont({area}, 0.99) area_p99,
          count(*) filter (where {annual} > 0 and {area} > 0) annual_rent_per_area_feasible_lines,
          quantile_cont(case when {annual} > 0 and {area} > 0 then {annual} / {area} end, 0.5) annual_rent_per_sqm_median
        from rent_contracts
        """,
    )
    summary.update(line_quality)
    summary.update(
        {
            "row_semantics": "One property line within an Ejari contract; identity is (contract_id, line_number).",
            "header_semantics": "Dates, registration type, contract_amount and annual_amount are contract-level values repeated on every property line.",
            "amount_semantics": "contract_amount is the contract-period total in AED; annual_amount is DLD's calculated annualized amount for non-12-month contracts.",
            "safe_counting_rule": "Count distinct contract_id for contracts and composite keys for property lines; never sum repeated header amounts across lines.",
            "community_amount_rule": "Use single-line/single-area contracts only unless a documented allocation method becomes available.",
            "registration_date": "not supplied",
            "cancellation_status": "not supplied",
            "unit_identifier": "not supplied; renewals/amendments cannot be linked authoritatively to a stable unit",
            "rooms": "not supplied as a clean numeric field; low-level subtype labels must not be parsed as authoritative room counts",
            "area_unit": "not explicit in the rent dictionary; rent-per-square-foot remains conditional until square-metre semantics are confirmed",
        }
    )
    periods = rows_dict(
        conn,
        f"""
        select strftime(start_date, '%Y-%m') period, count(*) contracts,
          count(*) filter (where try_cast(max_type as bigint) = 1) new_contracts,
          count(*) filter (where try_cast(max_type as bigint) = 2) renewed_contracts
        from phase3a_contract_rollup where start_date between date '1900-01-01' and date '2026-09-04'
        group by 1 order by 1
        """,
    )
    categories: list[dict[str, Any]] = []
    for dimension, field in [
        ("registration_type", "contract_reg_type_en"),
        ("business_property_type", "ejari_bus_property_type_en"),
        ("property_type", "ejari_property_type_en"),
        ("property_sub_type", "ejari_property_sub_type_en"),
        ("usage", "property_usage_en"),
        ("tenant_type", "tenant_type_en"),
    ]:
        categories.extend(
            {"dimension": dimension, "value": row["category_value"], "property_lines": row["records"]}
            for row in rows_dict(
                conn,
                f'''select coalesce(nullif(trim("{field}"), ''), '[missing]') category_value, count(*) records
                    from rent_contracts group by 1 order by records desc, category_value''',
            )
        )
    return summary, periods, categories


JOIN_SPECS = [
    ("transactions.procedure_id -> transaction_procedures.procedure_id", "transactions", "procedure_id", "transaction_procedures", "procedure_id", "numeric"),
    ("transaction_procedures.group_id -> transaction_groups.group_id", "transaction_procedures", "group_id", "transaction_groups", "group_id", "numeric"),
    ("transactions.trans_group_id -> transaction_groups.group_id", "transactions", "trans_group_id", "transaction_groups", "group_id", "numeric"),
    ("transactions.reg_type_id -> market_types.market_type_id", "transactions", "reg_type_id", "market_types", "market_type_id", "numeric"),
    ("transactions.area_id -> areas.area_id", "transactions", "area_id", "areas", "area_id", "numeric"),
    ("transactions.project_number -> projects.project_number", "transactions", "project_number", "projects", "project_number", "numeric"),
    ("rent_contracts.area_id -> areas.area_id", "rent_contracts", "area_id", "areas", "area_id", "numeric"),
    ("rent_contracts.project_number -> projects.project_number", "rent_contracts", "project_number", "projects", "project_number", "numeric"),
    ("buildings.project_id -> projects.project_id", "buildings", "project_id", "projects", "project_id", "numeric"),
    ("units.project_id -> projects.project_id", "units", "project_id", "projects", "project_id", "numeric"),
    ("land_registry.project_id -> projects.project_id", "land_registry", "project_id", "projects", "project_id", "numeric"),
    ("projects.developer_id -> developers.developer_id", "projects", "developer_id", "developers", "developer_id", "numeric"),
    ("projects.master_developer_id -> developers.developer_id", "projects", "master_developer_id", "developers", "developer_id", "numeric"),
]


def join_expression(field: str, mode: str) -> str:
    if mode == "numeric":
        return f'try_cast(nullif(trim("{field}"), \'\') as decimal(38, 8))'
    return f'lower(regexp_replace(trim("{field}"), \'\\s+\', \' \', \'g\'))'


def join_matrix(conn: duckdb.DuckDBPyConnection) -> list[dict[str, Any]]:
    result = []
    for name, left_table, left_field, right_table, right_field, mode in JOIN_SPECS:
        left = join_expression(left_field, mode)
        right = join_expression(right_field, mode)
        row = row_dict(
            conn,
            f"""
            with parent as (
              select {right} join_key, count(*) parent_rows
              from "{right_table}" where {right} is not null group by 1
            ), child as (
              select {left} join_key, count(*) child_rows
              from "{left_table}" group by 1
            )
            select sum(child_rows) child_rows,
              sum(child_rows) filter (where child.join_key is null) null_child_rows,
              sum(child_rows) filter (where child.join_key is not null) non_null_child_rows,
              sum(child_rows) filter (where parent.join_key is not null) matched_rows,
              sum(child_rows) filter (where child.join_key is not null and parent.join_key is null) unmatched_rows,
              sum(child_rows) filter (where parent_rows > 1) ambiguous_child_rows,
              count(*) filter (where parent_rows > 1) ambiguous_child_keys,
              (select count(*) from parent where parent_rows > 1) duplicate_parent_keys,
              (select max(parent_rows) from parent) max_parent_multiplicity
            from child left join parent using (join_key)
            """,
        )
        for field in ["child_rows", "null_child_rows", "non_null_child_rows", "matched_rows", "unmatched_rows", "ambiguous_child_rows", "ambiguous_child_keys", "duplicate_parent_keys", "max_parent_multiplicity"]:
            row[field] = row[field] or 0
        non_null = row["non_null_child_rows"] or 0
        matched = row["matched_rows"] or 0
        ambiguous = row["ambiguous_child_rows"] or 0
        row.update(
            {
                "relationship": name,
                "left_table": left_table,
                "left_field": left_field,
                "right_table": right_table,
                "right_field": right_field,
                "normalization": "trim + exact numeric cast" if mode == "numeric" else "trim + lowercase + whitespace collapse",
                "match_percent": round(100 * matched / non_null, 6) if non_null else None,
                "cardinality": "many-to-one" if not row["duplicate_parent_keys"] else "ambiguous parent key",
                "production_stability": "stable" if non_null and matched == non_null and not ambiguous else "partial/review",
                "authoritative_fuzzy_matching": False,
            }
        )
        if name.startswith("transactions.reg_type_id -> market_types"):
            row["production_stability"] = "invalid/incompatible domains"
            row["cardinality"] = "not a valid relationship"
            row["semantic_warning"] = "Transaction reg_type is Existing/Off-Plan with IDs 1/0; market_types is Primary/Secondary with IDs 1/2. The numeric overlap at 1 is coincidental and must not be joined."
        result.append(row)
    composite = row_dict(
        conn,
        """
        with parent as (
          select try_cast(procedure_id as decimal(38,8)) procedure_id,
            try_cast(group_id as decimal(38,8)) group_id, count(*) parent_rows
          from transaction_procedures group by 1,2
        ), child as (
          select try_cast(procedure_id as decimal(38,8)) procedure_id,
            try_cast(trans_group_id as decimal(38,8)) group_id, count(*) child_rows
          from transactions group by 1,2
        )
        select sum(child_rows) child_rows, 0 null_child_rows, sum(child_rows) non_null_child_rows,
          coalesce(sum(child_rows) filter (where parent.procedure_id is not null),0) matched_rows,
          coalesce(sum(child_rows) filter (where parent.procedure_id is null),0) unmatched_rows,
          coalesce(sum(child_rows) filter (where parent_rows > 1),0) ambiguous_child_rows,
          coalesce(count(*) filter (where parent_rows > 1),0) ambiguous_child_keys,
          (select count(*) from parent where parent_rows > 1) duplicate_parent_keys,
          (select max(parent_rows) from parent) max_parent_multiplicity
        from child left join parent using (procedure_id, group_id)
        """,
    )
    non_null = composite["non_null_child_rows"] or 0
    matched = composite["matched_rows"] or 0
    composite.update(
        {
            "relationship": "transactions.(procedure_id,trans_group_id) -> transaction_procedures.(procedure_id,group_id)",
            "left_table": "transactions",
            "left_field": "procedure_id+trans_group_id",
            "right_table": "transaction_procedures",
            "right_field": "procedure_id+group_id",
            "normalization": "trim + exact numeric cast on both fields",
            "match_percent": round(100 * matched / non_null, 6) if non_null else None,
            "cardinality": "many-to-one" if not composite["duplicate_parent_keys"] else "ambiguous parent key",
            "production_stability": "stable" if matched == non_null and not composite["ambiguous_child_rows"] else "partial/review",
            "authoritative_fuzzy_matching": False,
            "semantic_warning": "Use this composite relationship; procedure_id alone is ambiguous across Sales and Mortgages for six lease-to-own procedures.",
        }
    )
    result.insert(1, composite)
    return result


def privacy_registry(
    columns: list[dict[str, Any]],
    dictionaries: list[dict[str, Any]],
    phase0_classifications: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    descriptions = {(row["dataset"], row["field"]): row.get("description", "") for row in dictionaries}
    phase0 = {(row["dataset"], row["field"]): row["classification"] for row in phase0_classifications}
    rows = []
    for column in columns:
        dataset = column["dataset"]
        field = column["field"]
        text = f"{field} {descriptions.get((dataset, field), '')}".lower()
        if re.search(r"(?:phone|email|fax|nationality|gender|age_group|tenant_type|trainee)", text):
            classification = "personal/sensitive"
            publication = "prohibited"
            reason = "Contact, demographic or participant-characteristic field."
        elif field in {"transaction_id", "contract_id", "line_number", "property_id", "parent_property_id", "grandparent_property_id", "parcel_id", "unit_number", "building_number", "land_number", "land_sub_number", "pre_registration_number", "procedure_number", "participant_id", "request_id", "application_id"}:
            classification = "internal matching field"
            publication = "prohibited"
            reason = "Granular event, contract, participant or property locator."
        elif phase0.get((dataset, field)) == "aggregate-only":
            classification = "safe only after aggregation"
            publication = "aggregate-only"
            reason = "Fact-level measure or identifier from a granular source."
        elif phase0.get((dataset, field)) == "internal" or field == "load_timestamp" or field.startswith("_source"):
            classification = "internal matching field"
            publication = "prohibited"
            reason = "Pipeline lineage field."
        else:
            classification = "safe public dimension"
            publication = "public-dimension"
            reason = "Official categorical label or non-personal directory fact."
        rows.append(
            {
                "dataset": dataset,
                "field": field,
                "classification": classification,
                "publication": publication,
                "reason": reason,
            }
        )
    return rows


def enrich_inventory(
    inventory: list[dict[str, Any]],
    columns: list[dict[str, Any]],
    privacy: list[dict[str, Any]],
) -> None:
    columns_by_dataset: dict[str, list[dict[str, Any]]] = {}
    privacy_by_dataset: dict[str, list[dict[str, Any]]] = {}
    for row in columns:
        columns_by_dataset.setdefault(row["dataset"], []).append(row)
    for row in privacy:
        privacy_by_dataset.setdefault(row["dataset"], []).append(row)
    for dataset in inventory:
        name = dataset["dataset"]
        dataset["columns"] = [
            {
                "name": row["field"],
                "declared_type": row.get("declared_data_type"),
                "null_count": row["null_rows"],
                "null_percent": row["null_percent"],
                "approx_distinct": row["approx_distinct"],
            }
            for row in sorted(columns_by_dataset[name], key=lambda item: item["ordinal"])
        ]
        decisions = privacy_by_dataset[name]
        dataset["personally_identifying_or_sensitive_columns"] = sorted(
            row["field"] for row in decisions if row["classification"] == "personal/sensitive"
        )
        fields = [row["field"] for row in decisions]
        dataset["geographic_identifiers"] = sorted(
            field for field in fields if re.search(r"(?:area|zone|municipality|munc|parcel|land)_?(?:id|number|code)?$", field)
        )
        dataset["developer_project_property_identifiers"] = sorted(
            field
            for field in fields
            if re.search(r"(?:developer|project|property|building|unit|land|parcel).*(?:id|number)$", field)
        )


def metric_definitions() -> list[dict[str, Any]]:
    common_sale = "Unique transaction_id; transaction group 1 plus Phase 3A market_sale_candidate procedure allowlist; valid date; actual_worth > 0; exclude lease-to-own, administrative/ambiguous procedures and flagged outliers."
    common_rent = "Distinct contract_id; valid start/end; annual_amount > 0; use single-line/single-area contracts for geography-dependent amounts."
    return [
        metric("sale_transaction_count", "supported", common_sale, "count distinct transaction_id", "n/a", "transaction month", "area/property type/registration type", "Suppress cells below 10."),
        metric("sale_transaction_value", "conditionally supported", common_sale, "sum actual_worth", "n/a", "transaction month", "area/property type/registration type", "Currency needs authority confirmation; suppress cells below 10."),
        metric("median_sale_price", "conditionally supported", common_sale, "median actual_worth", "eligible sale events", "transaction month", "area/property type/registration type", "Currency needs confirmation; suppress cells below 10."),
        metric("median_price_per_sqft", "conditionally supported", common_sale + " procedure_area > 0.", "median(actual_worth / procedure_area / 10.7639104167)", "eligible sale events with square-metre area", "transaction month", "area/property type/registration type", "Winsorize/report extreme flags; currency confirmation required."),
        metric("apartment_vs_villa", "supported", common_sale, "count and medians computed separately", "eligible events in exact official subtype labels", "transaction month", "property subtype", "Never infer subtype from names."),
        metric("ready_vs_off_plan", "supported", common_sale, "count and medians by reg_type", "eligible events", "transaction month", "official registration type", "Label as existing/off-plan per source, not secondary/primary."),
        metric("primary_vs_secondary", "unsupported", "No seller/developer role is supplied.", "n/a", "n/a", "n/a", "n/a", "Do not use reg_type as a proxy."),
        metric("community_trends", "conditionally supported", common_sale, "monthly count/value/median", "eligible events", "transaction month", "matched area_id", "Minimum 10; confidence bands at 10/30/100 samples."),
        metric("project_trends", "conditionally supported", common_sale + " unique matched project_number.", "monthly count/value/median", "eligible project-linked events", "transaction month", "project_number", "Suppress unmatched/ambiguous and samples below 10."),
        metric("developer_activity", "conditionally supported", common_sale + " uniquely matched project and developer.", "count/value of linked project events", "eligible linked events", "transaction month", "developer", "Activity, not performance, ranking or sales by developer."),
        metric("period_change", "conditionally supported", "Two complete comparable periods using the same eligibility/version.", "current metric / previous metric - 1", "previous period metric", "month/quarter/year", "same dimension", "Suppress if either period fails threshold; flag partial latest month."),
        metric("rental_contract_count", "supported", common_rent, "count distinct contract_id", "n/a", "contract start month", "area/type/new-renewed", "Count header once; suppress cells below 10."),
        metric("median_annual_rent", "supported", common_rent, "median annual_amount", "eligible distinct contracts", "contract start month", "area/type/new-renewed", "Use DLD annual_amount; multi-area contracts excluded from local medians."),
        metric("rent_per_sqft", "conditionally supported", common_rent + " actual_area > 0.", "median(annual_amount / actual_area / 10.7639104167)", "eligible single-property contracts", "contract start month", "area/type", "Rent area unit needs authority confirmation."),
        metric("new_vs_renewed_rent", "supported", common_rent, "counts and medians by contract_reg_type_id", "eligible distinct contracts", "contract start month", "registration type", "No linking of successive contracts without stable unit identity."),
        metric("gross_rental_yield", "unsupported for Phase 3B publication", "Rent and sale populations lack a stable shared unit and directly comparable cohort.", "median annual rent", "median sale price", "aligned trailing period and property cohort", "community/type", "May be researched later; never call it investor return."),
    ]


def metric(name: str, confidence: str, eligible: str, numerator: str, denominator: str, date_basis: str, geography: str, rules: str) -> dict[str, Any]:
    return {
        "metric": name,
        "confidence": confidence,
        "eligible_record_definition": eligible,
        "numerator": numerator,
        "denominator": denominator,
        "required_columns": required_columns(name),
        "exclusions_and_deduplication": rules,
        "outlier_treatment": "Retain in private facts; exclude invalid/non-positive denominators; publish robust medians and flag beyond source-aware fences.",
        "minimum_sample_threshold": 10,
        "suppression_rule": "No public metric when sample_size < 10; 10-29 low confidence, 30-99 moderate, 100+ high, subject to quality flags.",
        "date_basis": date_basis,
        "geographic_grouping": geography,
        "known_limitations": rules,
    }


def required_columns(name: str) -> list[str]:
    if name.startswith("rent") or name in {"median_annual_rent", "new_vs_renewed_rent", "gross_rental_yield"}:
        return ["contract_id", "line_number", "contract_start_date", "annual_amount", "area_id", "actual_area", "contract_reg_type_id"]
    return ["transaction_id", "instance_date", "trans_group_id", "procedure_id", "actual_worth", "procedure_area", "area_id", "reg_type_id", "property_type_id", "property_sub_type_id"]


def quality_summary(
    columns: list[dict[str, Any]],
    transaction: dict[str, Any],
    rent: dict[str, Any],
    joins: list[dict[str, Any]],
) -> dict[str, Any]:
    high_null = sorted(
        [
            {"dataset": row["dataset"], "field": row["field"], "null_percent": row["null_percent"]}
            for row in columns
            if row["null_percent"] >= 50
        ],
        key=lambda item: (-item["null_percent"], item["dataset"], item["field"]),
    )
    return {
        "exact_duplicate_rows_all_datasets": 0,
        "high_null_fields_50pct_or_more": high_null,
        "invalid_transaction_dates": transaction["invalid_dates"],
        "transaction_dates_before_1900": transaction["dates_before_1900"],
        "transaction_dates_after_export": transaction["dates_after_export"],
        "negative_transaction_values": transaction["negative_worth"],
        "zero_transaction_values": transaction["zero_worth"],
        "rent_contract_line_key_duplicates": rent["duplicate_contract_line_keys"],
        "rent_declared_property_count_mismatches": rent["declared_property_count_mismatches"],
        "rent_inconsistent_header_contracts": rent["inconsistent_header_contracts"],
        "negative_rent_values": rent["negative_annual_amount"],
        "zero_rent_values": rent["zero_annual_amount"],
        "unresolved_join_rows": sum((row["unmatched_rows"] or 0) for row in joins),
        "ambiguous_join_rows": sum((row["ambiguous_child_rows"] or 0) for row in joins),
        "arabic_english_note": "Official bilingual labels are retained independently. Phase 0 found dictionary description swaps and differing distinct-name counts; do not translate or use one language as an authoritative key.",
        "unit_note": "Transaction procedure_area is documented in square metres and is empirically checked against meter_sale_price. Rent actual_area has no explicit unit in its dictionary and needs confirmation.",
        "schema_drift_note": "CSV chunks within each dataset share one materialized schema. Future snapshots must be fingerprinted before union or replacement.",
    }


def bilingual_quality(conn: duckdb.DuckDBPyConnection, inventory: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for dataset in inventory:
        fields = {column["name"] for column in dataset["columns"]}
        pairs = sorted((field, field[:-3] + "_ar") for field in fields if field.endswith("_en") and field[:-3] + "_ar" in fields)
        if not pairs:
            continue
        expressions = []
        for index, (english, arabic) in enumerate(pairs):
            en = f'nullif(trim("{english}"), \'\')'
            ar = f'nullif(trim("{arabic}"), \'\')'
            expressions.extend(
                [
                    f"count(*) filter (where {en} is not null and {ar} is not null) p{index}_both",
                    f"count(*) filter (where {en} is not null and {ar} is null) p{index}_english_only",
                    f"count(*) filter (where {en} is null and {ar} is not null) p{index}_arabic_only",
                    f"count(*) filter (where {en} is null and {ar} is null) p{index}_both_missing",
                    f"count(*) filter (where {en} = {ar}) p{index}_identical",
                    f"count(*) filter (where {en} is not null and regexp_matches({en}, '[ء-ي]')) p{index}_arabic_script_in_english",
                    f"count(*) filter (where {ar} is not null and not regexp_matches({ar}, '[ء-ي]')) p{index}_no_arabic_script_in_arabic",
                ]
            )
        row = row_dict(conn, f'''select {", ".join(expressions)} from "{dataset["table"]}"''')
        for index, (english, arabic) in enumerate(pairs):
            result.append(
                {
                    "dataset": dataset["dataset"],
                    "english_field": english,
                    "arabic_field": arabic,
                    "both_present": row[f"p{index}_both"],
                    "english_only": row[f"p{index}_english_only"],
                    "arabic_only": row[f"p{index}_arabic_only"],
                    "both_missing": row[f"p{index}_both_missing"],
                    "identical_non_null": row[f"p{index}_identical"],
                    "arabic_script_in_english": row[f"p{index}_arabic_script_in_english"],
                    "no_arabic_script_in_arabic": row[f"p{index}_no_arabic_script_in_arabic"],
                }
            )
    return result


def analytical_model() -> dict[str, Any]:
    return {
        "status": "design only; no migration created or applied",
        "private_layers": [
            {"table": "dld_source_manifests", "grain": "one source export", "purpose": "checksums, schema version, source dates and publication state"},
            {"table": "dld_source_files", "grain": "one file/chunk", "purpose": "file hash, row count, schema fingerprint and replacement lineage"},
            {"table": "dld_transaction_facts_private", "grain": "one unique transaction_id", "purpose": "typed transaction events; never browser-readable"},
            {"table": "dld_rent_contracts_private", "grain": "one contract_id", "purpose": "deduplicated header dates and amounts"},
            {"table": "dld_rent_contract_lines_private", "grain": "one (contract_id,line_number)", "purpose": "property-line dimensions without public raw access"},
            {"table": "dld_data_quality_results", "grain": "one run/check/dimension", "purpose": "reconciliation, invalid values, joins and suppression evidence"},
            {"table": "dld_market_publication_runs", "grain": "one atomic methodology/source publication", "purpose": "version, counts, rollback pointer and status"},
        ],
        "safe_dimensions": ["dld_dim_period", "dld_dim_community", "dld_dim_project", "dld_dim_developer", "dld_dim_property_type", "dld_dim_procedure", "dld_dim_market_segment"],
        "aggregate_tables": [
            {"table": "dld_market_sales_periodic", "grain": "period + geography + property type + existing/off-plan", "measures": ["sample_size", "transaction_count", "transaction_value", "median_price", "median_price_per_sqft"]},
            {"table": "dld_market_rents_periodic", "grain": "period + geography + property type + new/renewed", "measures": ["sample_size", "contract_count", "median_annual_rent", "median_rent_per_sqft"]},
        ],
        "public_views_or_functions": ["public_dld_market_overview", "public_dld_community_series", "public_dld_project_series", "public_dld_developer_activity", "search_public_dld_market_series"],
        "public_columns_only": ["dimension keys/slugs", "official safe names", "period", "metric value", "sample_size", "source_export_date", "methodology_version", "suppression_flag", "confidence_flag"],
        "forbidden_public_payloads": ["raw transactions", "raw contracts", "participant attributes", "unit/property/land identifiers", "internal join keys", "unsuppressed small cells"],
        "atomicity": "Stage a complete run, validate counts/schema/referential integrity/suppression, then switch one publication-run pointer in a transaction; any failure rolls back.",
    }


def update_architecture() -> dict[str, Any]:
    return {
        "snapshot_assessment": "Files sharing one export date and numbered chunk suffixes form full dataset snapshots, not independent increments.",
        "detection": ["Require a manifest with dataset, chunk order, byte size, SHA-256, row count, schema fingerprint and source export date.", "Reject duplicate package hashes and mixed export dates unless explicitly declared.", "Compare source keys and schema fingerprints with the prior accepted run before processing."],
        "stable_keys": {"transactions": ["transaction_id"], "rent_contract_lines": ["contract_id", "line_number"], "rent_contract_headers": ["contract_id"]},
        "idempotency": "Rebuilding the same package hash and methodology version produces the same private facts and aggregates; publication is a no-op if already complete.",
        "replacement_strategy": "Treat each accepted export as a replacement snapshot for the included dataset. Do not append chunk files or append a later full snapshot.",
        "corrections": "Diff stable keys and row hashes. Recompute affected periods plus the latest incomplete period; if semantics/schema change, rebuild all history under a new methodology version.",
        "late_arrivals": "Recompute every period containing inserted, removed or changed source keys and any trailing comparison windows that depend on those periods.",
        "freshness": ["source_export_date", "package_hash", "published_at", "methodology_version", "latest_complete_period", "quality_status"],
        "rollback": "Keep the prior immutable publication run addressable; pointer switch and all aggregate writes must share one transaction.",
    }


def storage_estimates(conn: duckdb.DuckDBPyConnection, source: Path, inventory: list[dict[str, Any]]) -> dict[str, Any]:
    sale_rows = scalar(conn, f"select count(*) from transactions where {market_sale_sql()}")
    transaction_month_area = scalar(
        conn,
        f"select count(*) from (select date_trunc('month', {date_expr('instance_date')}) period, area_id from transactions where {market_sale_sql()} and {date_expr('instance_date')} is not null group by 1, 2)",
    )
    rent_month_area = scalar(
        conn,
        "select count(*) from (select date_trunc('month', start_date) period, min_area area_id from phase3a_contract_rollup where start_date is not null and min_area is not distinct from max_area group by 1, 2)",
    )
    raw_bytes = sum(item["compressed_bytes"] for item in inventory)
    facts_uncompressed = int(sale_rows * 176 + 10_442_927 * 128 + 8_713_621 * 96)
    aggregate_rows = int(transaction_month_area + rent_month_area + 50_000)
    return {
        "source_files_bytes": raw_bytes,
        "phase0_duckdb_bytes": DEFAULT_DB.stat().st_size,
        "private_fact_model_uncompressed_estimate_bytes": facts_uncompressed,
        "private_fact_model_compressed_estimate_range_bytes": [int(facts_uncompressed * 0.25), int(facts_uncompressed * 0.55)],
        "estimated_public_aggregate_rows": aggregate_rows,
        "estimated_public_aggregate_bytes": aggregate_rows * 160,
        "assumptions": [
            "Normalized facts use typed dates/numerics and dictionary-encoded dimensions.",
            "Public series are precomputed by month and approved dimensions, not raw-row copies.",
            "Estimates include indexes/metadata only as a broad compression range and require measurement in Phase 3B.",
        ],
    }


def report_manifest(report_dir: Path) -> dict[str, Any]:
    files = []
    for path in sorted(report_dir.iterdir()):
        if path.is_file() and path.name not in {"manifest.json", "verification.json"}:
            files.append({"file": path.name, "bytes": path.stat().st_size, "sha256": sha256_file(path)})
    return {"schema_version": 1, "methodology_version": METHODOLOGY_VERSION, "files": files}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--database", type=Path, default=DEFAULT_DB)
    parser.add_argument("--reports", type=Path, default=DEFAULT_REPORTS)
    args = parser.parse_args()

    source = args.source.resolve()
    database = args.database.resolve()
    reports = args.reports.resolve()
    if not source.is_dir():
        raise SystemExit(f"Source folder not found: {source}")
    if not database.is_file():
        raise SystemExit(f"Phase 0 DuckDB not found: {database}")
    if len([p for p in source.iterdir() if p.is_dir()]) != EXPECTED_DATASETS:
        raise SystemExit("Expected exactly 25 source dataset folders")
    if len([p for p in source.rglob('*') if p.is_file()]) != EXPECTED_FILES:
        raise SystemExit("Expected exactly 62 source files")

    reports.mkdir(parents=True, exist_ok=True)
    phase0_datasets = read_json(PHASE0_REPORTS / "datasets.json")
    columns = read_json(PHASE0_REPORTS / "columns.json")
    dictionaries = read_json(PHASE0_REPORTS / "attribute_dictionary.json")
    keys = read_json(PHASE0_REPORTS / "candidate_keys.json")
    date_ranges = read_json(PHASE0_REPORTS / "date_ranges.json")
    phase0_classifications = read_json(PHASE0_REPORTS / "field_classification.json")
    inventory = source_inventory(source, phase0_datasets)
    add_candidate_keys(inventory, keys)
    add_date_coverage(inventory, date_ranges)

    privacy = privacy_registry(columns, dictionaries, phase0_classifications)
    enrich_inventory(inventory, columns, privacy)

    temp_directory = ROOT / "data" / "dld" / "local" / "phase3a-tmp"
    temp_directory.mkdir(parents=True, exist_ok=True)
    conn = duckdb.connect(str(database), read_only=True)
    conn.execute("set threads=1")
    conn.execute("set preserve_insertion_order=false")
    conn.execute("set memory_limit='4GB'")
    conn.execute(f"set temp_directory='{temp_directory.as_posix()}'")
    fingerprints = schema_fingerprints(conn, inventory)
    transaction, transaction_periods, transaction_categories = transaction_profile(conn)
    procedures = procedure_policy(conn)
    rent, rent_periods, rent_categories = rent_profile(conn)
    residential_index = residential_index_profile(conn)
    for dataset in inventory:
        if dataset["dataset"] == "Residential Sale Index":
            for coverage in dataset["date_coverage"]:
                if coverage["field"] == "first_date_of_month":
                    coverage.update(
                        {
                            "parsed_rows": residential_index["total_rows"],
                            "unparsed_rows": residential_index["invalid_period_keys"],
                            "minimum": residential_index["period_min"],
                            "maximum": residential_index["period_max"],
                            "phase3a_correction": "ISO date parsed as a month key; Phase 0's generic month parser did not recognize this representation.",
                        }
                    )
    joins = join_matrix(conn)
    bilingual = bilingual_quality(conn, inventory)
    metrics = metric_definitions()
    quality = quality_summary(columns, transaction, rent, joins)
    storage = storage_estimates(conn, source, inventory)
    conn.close()

    source_dates = sorted({row["source_export_date"] for row in inventory if row["source_export_date"]})
    metadata = {
        "schema_version": 1,
        "methodology_version": METHODOLOGY_VERSION,
        "source_path_resolved": str(source),
        "source_export_dates": source_dates,
        "latest_source_export_date": max(source_dates),
        "dataset_count": len(inventory),
        "file_count": sum(row["file_count"] for row in inventory),
        "source_row_count": sum(row["row_count"] for row in inventory),
        "raw_rows_emitted": 0,
        "database_opened_read_only": True,
    }
    write_json(reports / "metadata.json", metadata)
    write_json(reports / "inventory.json", inventory)
    write_csv(
        reports / "inventory.csv",
        [
            {key: row[key] for key in ["dataset", "table", "dataset_kind", "file_count", "compressed_bytes", "uncompressed_bytes", "row_count", "column_count", "source_export_date"]}
            for row in inventory
        ],
    )
    write_json(reports / "schema_fingerprints.json", fingerprints)
    write_json(reports / "bilingual_quality.json", bilingual)
    write_csv(reports / "bilingual_quality.csv", bilingual)
    write_json(reports / "transaction_profile.json", transaction)
    write_csv(reports / "transaction_periods.csv", transaction_periods)
    write_csv(reports / "transaction_categories.csv", transaction_categories)
    write_json(reports / "procedure_policy.json", procedures)
    write_csv(reports / "procedure_policy.csv", procedures)
    write_json(reports / "rent_profile.json", rent)
    write_csv(reports / "rent_periods.csv", rent_periods)
    write_csv(reports / "rent_categories.csv", rent_categories)
    write_json(reports / "residential_index_profile.json", residential_index)
    write_json(reports / "join_matrix.json", joins)
    write_csv(reports / "join_matrix.csv", joins)
    write_json(reports / "privacy_registry.json", privacy)
    write_csv(reports / "privacy_registry.csv", privacy)
    write_json(reports / "metric_definitions.json", metrics)
    write_csv(reports / "metric_definitions.csv", metrics)
    write_json(reports / "quality_summary.json", quality)
    write_json(reports / "storage_estimates.json", storage)
    write_json(reports / "analytical_model.json", analytical_model())
    write_json(reports / "update_architecture.json", update_architecture())
    write_json(reports / "manifest.json", report_manifest(reports))
    print(json.dumps({"reports": str(reports), **metadata}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
